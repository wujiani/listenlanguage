import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import requests

import openpyxl
import os
from datetime import datetime


def login_and_get_cookies(browser):
    # 登录
    webwait = WebDriverWait(browser, 6)
    step_1 = webwait.until(
        EC.presence_of_element_located((By.XPATH, '/html/body/main/header/div/nav/div/div/ul/li[6]/a')))
    step_1.click()

    # 输入邮箱
    email_input = browser.find_element(By.XPATH, '/html/body/main/section/div/div/div/div/form/div[1]/input')
    email_input.send_keys('jiani.wu@studenti.unipd.it')

    # 输入密码
    password_input = browser.find_element(By.XPATH, '/html/body/main/section/div/div/div/div/form/div[2]/input')
    password_input.send_keys('Keita5+1keita')

    # 点击登录按钮
    stay_login_button = browser.find_element(By.CLASS_NAME, 'fa')
    stay_login_button.click()

    # 关闭Cookie提示框
    cookie_consent = browser.find_element(By.XPATH, '/html/body/div[2]/div[2]')
    cookie_consent.click()

    login_button = browser.find_element(By.XPATH, '/html/body/main/section/div/div/div/div/form/div[3]/button')
    login_button.click()
    time.sleep(5)

    # 等待登录成功并获取cookie
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.ID, 'id_text')))

    cookies = browser.get_cookies()
    # 将 cookie 列表转换为一个字典
    cookie_dict = {}
    for cookie in cookies:
        cookie_dict[cookie['name']] = cookie['value']
    return cookie_dict

def generate_and_download_audio(sheet, browser, text, language_value, voice_value, cookie_dict, i, mp3_path):
    browser.refresh()
    # 输入文字
    input_text = browser.find_element(By.ID, 'id_text')
    input_text.clear()
    input_text.send_keys(text)

    # 选择语言和声音
    sel_language = Select(browser.find_element(By.ID, 'id_language'))
    sel_language.select_by_value(language_value)
    sel_voice = Select(browser.find_element(By.ID, 'id_voice'))
    sel_voice.select_by_value(voice_value)

    # 提交表单
    submit_button = browser.find_element(By.ID, 'submit')
    submit_button.click()

    # 等待下载按钮出现
    WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.ID, 'download-button')))

    # 获取下载链接
    download_button = browser.find_element(By.ID, 'download-button')
    download_link = download_button.get_attribute("href")

    # 下载文件并保存
    download_and_save_file(sheet, text, download_link, cookie_dict, i, mp3_path)

    time.sleep(5)

def download_and_save_file(sheet, filename, download_link, cookie_dict, i, mp3_path):
    print(download_link)
    response = requests.get(download_link, cookies=cookie_dict)
    if response.status_code == 200:
        print("down",download_link)
        file_index = download_link.split('/')[-2]
        file_name = file_index+'.mp3'
        print(file_name)
        mp3_path_concrete = os.path.join(mp3_path, file_name)
        print(mp3_path_concrete)
        os.makedirs(os.path.dirname(mp3_path_concrete), exist_ok=True)
        print(mp3_path_concrete)
        with open(mp3_path_concrete, "wb") as file:
            file.write(response.content)
        print(f"{filename} 文件下载成功")
        sheet.cell(row=i, column=3, value=file_index)
        print("i")
    else:
        print(f"{filename} 文件下载失败")

# 创建浏览器实例
chrome_options = webdriver.ChromeOptions()
browser = webdriver.Chrome(options=chrome_options)

# 打开网页
browser.get('https://ondoku3.com/zh-hans/')

# 登录并获取cookie
cookie_dict = login_and_get_cookies(browser)


# 打开Excel文件
def get_file_paths(date_str):
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')

    year = date_obj.strftime('%Y')
    month = date_obj.strftime('%Y-%m')
    day = date_obj.strftime('%Y-%m-%d')

    xlsx_path = os.path.join('.\\context', year, month, day, day + '.xlsx')
    year_month = date_str[:7]  # 提取年和月部分，例如：'2024-02'
    mp3_path = f"./context/{date_str[:4]}/{year_month}/{date_str}/"

    return xlsx_path, mp3_path

# 使用示例
###############################################
date_str = '2024-02-21'
###############################################
xlsx_path, mp3_path = get_file_paths(date_str)

# excel_file = 'C:/Users/19wuj/PycharmProjects/listenlanguage2024-02-16.xlsx'
workbook = openpyxl.load_workbook(xlsx_path)

#
# from openpyxl import load_workbook
#
# # 打开 Excel 文件
# workbook = load_workbook(excel_file)

languages = {
            # 'jp':('ja-JP', 'ja-JP-Takumi-NTTS'),
             'it':('it-IT', 'it-IT-BenignoNeural'),
             # 'kr':('ko-KR', 'ko-KR-JennyMultilingualV2Neural')
            }

# 选择工作表
for language in languages:
    try:
        sheet = workbook[language]
    except KeyError:
        continue
    language_value = languages[language][0]
    voice_value = languages[language][1]

    # 从第一列中读取所有单元格的值，并存储在一个列表中
    phrases = [cell.value for cell in sheet['A'] if cell.value]
    # # 待生成和下载的文字列表
    # texts = phrases

    # 写入数据到第三列
    for i, text in enumerate(phrases, start=1):
        print("i",i)
        generate_and_download_audio(sheet, browser, text, language_value, voice_value, cookie_dict,i, mp3_path)


# 保存更改
workbook.save(xlsx_path)


# 关闭浏览器
browser.quit()







